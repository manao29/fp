"""Excel / PDF 解析工具 — 批量开票资料识别"""
from __future__ import annotations

import logging
import uuid
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ── Excel 解析 ──────────────────────────────────────────────

# 常见列名映射: (关键词列表, 目标字段)
EXCEL_COLUMN_MAPPINGS = [
    (["销售方公司", "销售方", "卖方公司", "开票方", "seller"], "seller_company_name"),
    (["销售方税号", "卖方税号"], "seller_tax_id"),
    (["公司名称", "抬头", "购买方", "买方名称", "客户名称", "名称", "company"], "company_name"),
    (["税号", "纳税人识别号", "tax_id", "tax"], "tax_id"),
    (["项目", "商品名称", "服务名称", "品名", "货物名称", "item", "product"], "item_name"),
    (["数量", "quantity", "qty", "count"], "quantity"),
    (["单位", "unit", "计量单位"], "unit"),
    (["金额", "总金额", "含税金额", "价税合计", "amount", "total"], "amount"),
    (["税率", "tax_rate", "rate"], "tax_rate"),
    (["发票类型", "invoice_type", "type"], "invoice_type"),
    (["邮箱", "email", "mail"], "buyer_email"),
    (["订单号", "order_no", "order"], "order_no"),
    (["备注", "remark", "说明"], "remark"),
    (["身份证号", "身份证", "id_card"], "id_card_no"),
    (["地址", "address", "addr"], "address"),
    (["电话", "phone", "tel"], "phone"),
    (["开户行", "bank_name", "bank"], "bank_name"),
    (["银行账号", "bank_account", "account"], "bank_account"),
]


def normalize_header(text: str) -> str:
    """标准化表头文字"""
    return str(text or "").strip().replace(" ", "").replace("\n", "").lower()


def map_columns_to_fields(headers: list[str]) -> dict[int, str]:
    """将 Excel 列名映射到标准字段"""
    mapping: dict[int, str] = {}
    for col_idx, header in enumerate(headers):
        norm = normalize_header(header)
        if not norm:
            continue
        for keywords, field in EXCEL_COLUMN_MAPPINGS:
            target_keywords = [normalize_header(k) for k in keywords]
            if norm in target_keywords:
                mapping[col_idx] = field
                break
    return mapping


def parse_excel_rows(file_path: str) -> list[dict[str, str]]:
    """
    解析 Excel 文件中的开票资料，每行返回一条 dict。
    支持 .xlsx 和 .xls 格式。
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    rows: list[list[str]] = []

    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell or "").strip() for cell in row])
        wb.close()

    elif suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            rows.append([str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)])

    else:
        return []

    if len(rows) < 2:
        return []

    # 第一行: 表头
    headers = rows[0]
    col_mapping = map_columns_to_fields(headers)

    results: list[dict[str, str]] = []
    batch_id = uuid.uuid4().hex[:12]
    data_rows = rows[1:]

    for row_idx, row in enumerate(data_rows):
        record: dict[str, str] = {"input_type": "excel"}
        is_empty = True

        for col_idx, field in col_mapping.items():
            if col_idx < len(row):
                value = str(row[col_idx]).strip()
                if value:
                    record[field] = value
                    is_empty = False

        if is_empty:
            continue

        # 个人票判断
        if not record.get("subject_type"):
            company = record.get("company_name", "")
            if company and ("有限" in company or "股份" in company or "工商" in company):
                record["subject_type"] = "company"
            elif record.get("id_card_no") or len(company) <= 4:
                record["subject_type"] = "person"

        # 数量格式化
        qty = record.get("quantity", "")
        if qty:
            try:
                record["quantity"] = str(int(float(qty))) if float(qty) == int(float(qty)) else str(float(qty))
            except ValueError:
                pass

        # 金额格式化
        amt = record.get("amount", "")
        if amt:
            amt = amt.replace(",", "").replace("¥", "").replace("￥", "").replace("元", "").strip()
            try:
                record["amount"] = str(float(amt))
            except ValueError:
                pass

        record["batch_id"] = batch_id
        record["batch_index"] = str(row_idx + 1)
        record["batch_total"] = str(len(data_rows))

        results.append(record)

    return results


# ── PDF 解析 ────────────────────────────────────────────────

def parse_pdf_text(file_path: str) -> str:
    """从 PDF 文件中提取文本"""
    path = Path(file_path)
    if not path.exists():
        return ""

    text_parts: list[str] = []

    # 优先 pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        if text_parts:
            return "\n".join(text_parts)
    except Exception as e:
        logger.debug("pdfplumber failed: %s", e)

    # 回退 PyPDF2
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text_parts.append(t)
    except Exception as e:
        logger.warning("PyPDF2 failed: %s", e)

    return "\n".join(text_parts)
